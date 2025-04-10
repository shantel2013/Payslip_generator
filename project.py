# print("hello world")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import fonts , alignment
import os
from fpdf import FPDF
import smtplib 
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText  
from email.mime.base import MIMEBase
from email import encoders

name = "PAYSLIP COMPANY"
employee_data = pd.read_excel('payslip.xlsx', engine='openpyxl')
employee_data.columns = employee_data.columns.str.strip()
print(employee_data.columns)

data = {
    "EMPLOYEE NAME": ["KYLE ", "KUDZI", "TYNOE ", "ANNA ","WESLEY ", ],
    "EMPLOYEE ID":["D766",  "D455" , "D677", "D966" , "D811"],
    "EMAIL" : ["tinomudaishekutama2004@gmail.com","vincent@uncommon.org","chigurilyncia@gmail.com","chakurungamaelsie1@gmail.com","nyabungawesley@gmail.com"],      
    "BASIC SALARY": [500 , 600 , 3000 , 4000 , 1000],
    "ALLOWANCE": [300 , 100 , 800 , 2000 , 900],
    "DEDUCTION" : [150 ,300 ,1000, 205 ,500],
    "NET SALARY" : [250 , 350 , 2000, 1500 , 450]   
    
    }

df = pd.DataFrame(data)

excel_file = "payslip.xlsx"

if os.path.exists(excel_file):
     print(f"the file '{excel_file}'already exists.please choose a defferent name or delete the exists file ")
else:
    df.to_excel(excel_file , index=False)
    print(f"data exported to '{excel_file}'  was successfuly")



excel_data = pd.read_excel("payslip.xlsx")

for index in range (len(df)):
    employee_infor = excel_data.loc[index]
    print(f"employee infor for index {index}:\n{employee_infor}:\n")
      
# function

def generate_pdf(employee):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt = f"Payslip for {employee['EMPLOYEE NAME']}", ln = True, align = 'C')
    pdf.cell(200, 10, txt = f"Employee ID: {employee['EMPLOYEE ID']}", ln = True, align = 'C')
    pdf.cell(200, 10, txt = f"Basic Salary:$ {employee['BASIC SALARY']}", ln = True, align = 'C')
    pdf.cell(200, 10, txt = f"Allowances:$ {employee['ALLOWANCE']}", ln = True, align = 'C')
    pdf.cell(200, 10, txt = f"Deductions: ${employee['DEDUCTION']}", ln = True, align = 'C')
    pdf.cell(200, 10, txt = f"Net Salary: ${employee['NET SALARY']}", ln = True, align = 'C')

    # Save PDF
    pdf_file_path = f"{employee['EMPLOYEE NAME'].replace(' ', '_')}_payslip.pdf"
    pdf.output(pdf_file_path)
    
### Generate payslips for each emplo

for index, employee in df.iterrows():
    generate_pdf(employee)
    
# Create the email
def send_email_with_payslip(employee):
    from_email = "tinomudaishekutama2004@gmail.com"
    to_email = employee["EMAIL"]
    subject = f"Payslip for {employee['EMPLOYEE NAME']}"
    body = f"Dear {employee['EMPLOYEE NAME']},\n\nPlease find attached your payslip for this month.\n\nBest regards,\n{name}"
    
    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    pdf_file = f"{employee['EMPLOYEE NAME'].replace(' ', '_')}_payslip.pdf"
    with open(pdf_file, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={pdf_file}')
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(from_email,"rgdeebyzbaaofsiu")
            server.send_message(msg)
            print(f"Payslip sent to {to_email}")
    except Exception as e:
        print(f"Failed to send email to {to_email}: {e}")

for index, employee in df.iterrows():
    generate_pdf(employee)
    send_email_with_payslip(employee)
